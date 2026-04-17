import os
import glob
import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime
import random
import requests
import base64
import time

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def get_image_caption(image_path, api_key):
    if not api_key:
        return "ไม่มี API Key ระบุ"
    
    base64_image = encode_image(image_path)
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": "typhoon-ocr", 
        "messages": [
            {
                "role": "system",
                "content": "คุณคือวิศวกรประเมินหน้างานก่อสร้าง ให้จัดหมวดหมู่งานจากรูปภาพ คัดลอกเฉพาะข้อความหมวดหมู่ต่อไปนี้ไปตอบเท่านั้น ห้ามใส่ตัวเลข ห้ามพิมพ์คำอื่นเพิ่ม: งานโครงผนัง, งานตกแต่งภายใน, งานฝ้าเพดาน, งานระบบปรับอากาศ, งานระบบไฟฟ้า, งานทำความสะอาด, งานสี, งานฉาบผนัง, งานสุขภัณฑ์, งานเตรียมพื้นที่ เตรียมวัสดุ, งานแก้ไขงานก่อสร้าง ถ้ารูปดูยากหรือไม่แน่ใจให้ตอบ งานเตรียมพื้นที่ เตรียมวัสดุ"
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "เลือก 1 หมวดหมู่ที่อธิบายภาพนี้ได้ดีที่สุด พิมพ์ตอบแค่ชื่อหมวดหมู่นั้น"
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 1024,
        "temperature": 0.4
    }
    
    for attempt in range(3):
        try:
            response = requests.post("https://api.opentyphoon.ai/v1/chat/completions", headers=headers, json=payload, timeout=30)
            if response.status_code == 200:
                data = response.json()
                raw_result = data["choices"][0]["message"]["content"].strip()
                valid_types = [
                    "งานโครงผนัง", "งานตกแต่งภายใน", "งานฝ้าเพดาน", "งานระบบปรับอากาศ",
                    "งานระบบไฟฟ้า", "งานทำความสะอาด", "งานสี", "งานฉาบผนัง",
                    "งานสุขภัณฑ์", "งานเตรียมพื้นที่ เตรียมวัสดุ", "งานแก้ไขงานก่อสร้าง"
                ]
                result = "งานเตรียมพื้นที่ เตรียมวัสดุ"
                for vt in valid_types:
                    if vt in raw_result:
                        result = vt
                        break
                return result
            elif response.status_code == 429:
                time.sleep(5)
            else:
                return "งานเตรียมพื้นที่ เตรียมวัสดุ"
        except Exception:
            time.sleep(2)
            
    return "งานเตรียมพื้นที่ เตรียมวัสดุ"

def get_daily_summary(captions, api_key):
    if not api_key:
        return "ไม่มีข้อมูลการทำงานในวันนี้"
        
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    captions_text = ", ".join(captions)
    payload = {
        "model": "typhoon-v2.1-12b-instruct",
        "messages": [
            {
                "role": "system",
                "content": "คุณคือวิศวกรคุมงานก่อสร้าง สรุปการทำงานประจำวันจากรายการงานที่ทำแบบสั้นๆ กระชับ ได้ใจความ (ความยาวไม่เกิน 1-2 บรรทัด)"
            },
            {
                "role": "user",
                "content": f"รูปภาพของรายงานในวันนี้แสดงถึงการทำงานดังนี้: {captions_text} \nช่วยสรุปสั้นๆ ว่าจากรูปภาพเหล่านี้ ทำงานอะไรบ้าง"
            }
        ],
        "max_tokens": 1024,
        "temperature": 0.5
    }
    
    for attempt in range(3):
        try:
            response = requests.post("https://api.opentyphoon.ai/v1/chat/completions", headers=headers, json=payload, timeout=30)
            if response.status_code == 200:
                data = response.json()
                return data["choices"][0]["message"]["content"].strip()
            elif response.status_code == 429:
                time.sleep(5)
            else:
                return f"งานที่ดำเนินการวันนี้หลักๆ ได้แก่: {captions_text}"
        except Exception:
            time.sleep(2)
            
    return f"งานที่ดำเนินการวันนี้หลักๆ ได้แก่: {captions_text}"

def generate_report(project_dir, report_date_str, api_key):
    template_path = os.path.join(project_dir, "ข้อมูลพื้นฐาน", "REPORT_template_00.xlsx")
    picture_dir = os.path.join(project_dir, "รูปถ่ายจากพื้นที่ก่อสร้าง", "รูปถ่ายยังไม่แยก")
    output_dir = os.path.join(project_dir, "รายงานประจำวัน")
    project_info_path = os.path.join(project_dir, "ข้อมูลพื้นฐาน", "ชื่อโครงการ.xlsx")
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Load Project Name
    project_name = "ชื่อโครงการ"
    if os.path.exists(project_info_path):
        try:
            wb_info = load_workbook(project_info_path)
            ws_info = wb_info.active
            project_name = ws_info['B1'].value or "ชื่อโครงการ"
        except Exception:
            pass

    date_obj = datetime.strptime(report_date_str, "%Y-%m-%d")
    thai_date = f"{date_obj.day}/{date_obj.month:02d}/{date_obj.year + 543}"
    
    output_file = os.path.join(output_dir, f"DailyReport_{report_date_str}.xlsx")
    
    image_slots = [
        {'cell': 'E66', 'text': 'E67'},
        {'cell': 'Z66', 'text': 'Z67'},
        {'cell': 'E69', 'text': 'E70'},
        {'cell': 'Z69', 'text': 'Z70'},
        {'cell': 'E72', 'text': 'E73'},
        {'cell': 'Z72', 'text': 'Z73'},
    ]
    
    wb = load_workbook(template_path)
    ws = wb.active 

    # Cell Mapping based on template analysis
    ws['C6'] = project_name
    ws['AK7'] = thai_date 
    ws['Y10'] = "08:00 - 17:00" # Default or customizable

    # Find images for the day (if folder exists or directly in unseparated)
    # The script used folders by date, user said they put photos in unseparated.
    # We can try to filter by date from metadata or use a subfolder if user provided it.
    
    # If unseparated is the folder, we look for images there.
    images = glob.glob(os.path.join(picture_dir, "*.jpg")) + glob.glob(os.path.join(picture_dir, "*.png"))
    
    # Filter by date if possible, but user said "I will put photos then command you"
    # So we assume the current photos in that folder are for the report.
    
    daily_captions = []
    for idx in range(min(6, len(images))):
        img_path = images[idx]
        try:
             img = OpenpyxlImage(img_path)
             slot = image_slots[idx]
             
             # Resize to fit slot (roughly)
             img.width = 500
             img.height = 375
             
             ws.add_image(img, slot['cell'])
             
             print(f"  Processing image {idx+1}/{len(images)}...")
             caption = get_image_caption(img_path, api_key)
             ws[slot['text']] = caption
             daily_captions.append(caption)
             time.sleep(1)
        except Exception as e:
             print(f"  Warning: Could not insert {img_path}: {e}")

    if daily_captions:
        summary = get_daily_summary(daily_captions, api_key)
        ws['N34'] = summary
    else:
        ws['N34'] = "ไม่มีภาพถ่ายของงานในวันนี้"

    wb.save(output_file)
    print(f"Report saved: {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--project_dir", required=True)
    parser.add_argument("--date", required=True)
    parser.add_argument("--api_key", required=True)
    args = parser.parse_args()
    
    generate_report(args.project_dir, args.date, args.api_key)
