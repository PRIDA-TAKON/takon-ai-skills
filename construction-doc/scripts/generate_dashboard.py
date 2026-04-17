import os
import glob
import argparse
from openpyxl import load_workbook
import json

def generate_dashboard(project_dir):
    boq_path = os.path.join(project_dir, "ข้อมูลพื้นฐาน", "BOQ.xlsx")
    reports_dir = os.path.join(project_dir, "รายงานประจำวัน")
    project_info_path = os.path.join(project_dir, "ข้อมูลพื้นฐาน", "ชื่อโครงการ.xlsx")
    
    # Load Project Name
    project_name = os.path.basename(project_dir)
    if os.path.exists(project_info_path):
        try:
            wb_info = load_workbook(project_info_path)
            ws_info = wb_info.active
            project_name = ws_info['B1'].value or project_name
        except Exception:
            pass

    # Load BOQ Items
    boq_items = []
    if os.path.exists(boq_path):
        wb_boq = load_workbook(boq_path, data_only=True)
        ws_boq = wb_boq.active
        # Assuming Item Description is Column C, starting Row 2
        for row in range(2, ws_boq.max_row + 1):
            item_desc = ws_boq[f'C{row}'].value
            if item_desc:
                boq_items.append({
                    "id": ws_boq[f'A{row}'].value,
                    "desc": item_desc,
                    "progress": 0 # Default
                })

    # Read latest progress from reports
    # This is tricky because we don't know where the user puts the % in the template yet.
    # But we can look for the most recent report.
    report_files = glob.glob(os.path.join(reports_dir, "DailyReport_*.xlsx"))
    report_files.sort(reverse=True)
    
    if report_files:
        latest_report = report_files[0]
        try:
            wb_rep = load_workbook(latest_report, data_only=True)
            ws_rep = wb_rep.active
            # User said: "Work Items/Record starts at Row 32 (Location in C32, Work Type in M32)"
            # Let's assume progress % is in another column, e.g., Column Q or similar.
            # For now, let's just mock some progress or try to find it.
            # I will assume there's a mapping logic.
            pass
        except Exception:
            pass

    # Simple HTML Template
    html_template = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Dashboard - {project_name}</title>
    <meta charset="utf-8">
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f4f7f6; margin: 0; padding: 20px; }}
        .container {{ max-width: 1000px; margin: auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; border-bottom: 2px solid #eee; padding-bottom: 10px; }}
        .work-item {{ margin-bottom: 15px; }}
        .label {{ font-weight: bold; margin-bottom: 5px; display: block; }}
        .progress-bg {{ background: #e0e0e0; border-radius: 10px; height: 20px; overflow: hidden; }}
        .progress-bar {{ background: #4caf50; height: 100%; transition: width 0.5s; }}
        .percent {{ float: right; font-size: 0.9em; color: #666; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Dashboard: {project_name}</h1>
        <div id="items">
            {''.join([f'''
            <div class="work-item">
                <span class="label">{item['desc']} <span class="percent">{item['progress']}%</span></span>
                <div class="progress-bg">
                    <div class="progress-bar" style="width: {item['progress']}%"></div>
                </div>
            </div>
            ''' for item in boq_items])}
        </div>
    </div>
</body>
</html>
    """
    
    output_path = os.path.join(project_dir, "แผนงานและความคืบหน้า", "dashboard.html")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"Dashboard generated: {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--project_dir", required=True)
    args = parser.parse_args()
    generate_dashboard(args.project_dir)
